import os
import re
import subprocess
import pandas as pd
import datetime
import shutil
import requests
from openpyxl import load_workbook


def close_excel_if_open(filepath=None):
    abs_path = os.path.abspath(filepath) if filepath else None
    try:
        result = subprocess.run(
            ["powershell", "-Command",
             f'Get-Process EXCEL -ErrorAction SilentlyContinue | Where-Object {{ $_.MainWindowTitle -ne "" }} | ForEach-Object {{ $_.CloseMainWindow() }}; '
             f'Start-Sleep -Seconds 1; '
             f'Get-Process EXCEL -ErrorAction SilentlyContinue | Stop-Process -Force'],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode == 0:
            print("Closed Excel processes to release file lock.")
            return True
    except Exception as e:
        print(f"Could not close Excel gracefully: {e}")
    try:
        subprocess.run(["taskkill", "/f", "/im", "EXCEL.EXE"],
                       capture_output=True, timeout=5)
        print("Force-closed Excel processes.")
        return True
    except Exception as e:
        print(f"Could not force-close Excel: {e}")
        return False


def _normalize_date(d):
    if not d or d == "0001-01-01":
        return ""
    d = str(d).strip()
    if re.match(r"^\d{1,2}\s+[A-Za-z]{3}\s+\d{4}$", d):
        return d
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", d)
    if m:
        y, mo, day = m.groups()
        month_map = {"01":"Jan","02":"Feb","03":"Mar","04":"Apr","05":"May","06":"Jun",
                     "07":"Jul","08":"Aug","09":"Sep","10":"Oct","11":"Nov","12":"Dec"}
        return f"{int(day):d} {month_map.get(mo, mo)} {y}"
    return d


def export_combined(batting, bowling):
    os.makedirs("reports", exist_ok=True)

    df_b = pd.DataFrame(batting)
    df_bo = pd.DataFrame(bowling)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    latest_path = "reports/AllTeams_Stats_latest.xlsx"
    backup_path = f"reports/AllTeams_Stats_{timestamp}.xlsx"

    def _to_sheet_dict():
        out = {}
        if not df_b.empty:
            df_b_out = df_b.rename(columns={
                "CompetitionID": "Competition",
                "TeamName": "Team",
                "PlayerName": "Player",
                "HighestScore": "Highest Score",
                "BattingAverage": "Average",
                "StrikeRate": "Strike Rate",
                "Hundreds": "100s",
                "Fifties": "50s",
            })
            out["Batting"] = df_b_out
        else:
            out["Batting"] = pd.DataFrame()

        if not df_bo.empty:
            out["Bowling"] = df_bo
        else:
            out["Bowling"] = pd.DataFrame()
        return out

    sheets = _to_sheet_dict()

    # Try xlwings first to support updating while Excel is open
    try:
        import xlwings as xw

        def _save_with_xlwings(path, sheets_dict):
            app = xw.App(visible=False)
            book = None
            try:
                # Try to open for write access; some Excel versions accept ReadOnly arg
                if os.path.exists(path):
                    try:
                        book = app.books.open(path, ReadOnly=False)
                    except Exception:
                        # fallback to default open
                        book = app.books.open(path)
                else:
                    book = app.books.add()
                    # ensure workbook has a sensible name before saving
                    book.save(path)

                # write each sheet
                for name, df in sheets_dict.items():
                    if name in [s.name for s in book.sheets]:
                        sht = book.sheets[name]
                    else:
                        sht = book.sheets.add(name)
                    try:
                        sht.clear()
                    except Exception:
                        pass
                    # xlwings raises AssertionError when writing empty frames with zero columns
                    if df is None or (hasattr(df, 'empty') and df.empty) or (hasattr(df, 'shape') and df.shape[1] == 0):
                        # leave sheet cleared
                        continue
                    sht.range("A1").options(index=False).value = df

                book.save(path)
                book.close()
                app.quit()
                return True
            except Exception:
                try:
                    if book:
                        book.close()
                except Exception:
                    pass
                try:
                    app.quit()
                except Exception:
                    pass
                raise

        _save_with_xlwings(latest_path, sheets)
        print("Latest Excel updated:", latest_path)
        shutil.copyfile(latest_path, backup_path)
        print("Backup Excel saved:", backup_path)
    except Exception as e:
        import traceback
        print("xlwings error in export_combined:")
        print(traceback.format_exc())
        # fallback to openpyxl writer (handles normal case)
        try:
            with pd.ExcelWriter(latest_path, engine="openpyxl") as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            print("Latest Excel updated:", latest_path)

            with pd.ExcelWriter(backup_path, engine="openpyxl") as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            print("Backup Excel saved:", backup_path)
        except PermissionError:
            alt = f"reports/AllTeams_Stats_{timestamp}_alt.xlsx"
            with pd.ExcelWriter(alt, engine="openpyxl") as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"Original latest file locked. Saved alternate workbook to {alt}")


def export_google_sheet_workbook(sheet_dict):
    os.makedirs("reports", exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    latest_path = "reports/AllTeams_Stats_latest.xlsx"
    backup_path = f"reports/AllTeams_Stats_{timestamp}.xlsx"

    # Try xlwings first so updates can occur while workbook is open in Excel
    try:
        import xlwings as xw

        app = xw.App(visible=False)
        book = None
        try:
            if os.path.exists(latest_path):
                book = app.books.open(latest_path)
            else:
                book = app.books.add()
                book.save(latest_path)

            for sheet_name, df in sheet_dict.items():
                if sheet_name in [s.name for s in book.sheets]:
                    sht = book.sheets[sheet_name]
                else:
                    sht = book.sheets.add(sheet_name)
                sht.clear()
                if df is None or (hasattr(df, 'empty') and df.empty) or (hasattr(df, 'shape') and df.shape[1] == 0):
                    continue
                sht.range("A1").options(index=False).value = df

            book.save(latest_path)
            book.close()
            app.quit()
            print("Latest Excel updated:", latest_path)
            shutil.copyfile(latest_path, backup_path)
            print("Backup Excel saved:", backup_path)
            return
        except Exception as e:
            import traceback
            try:
                if book:
                    book.close()
            except Exception:
                pass
            try:
                app.quit()
            except Exception:
                pass
            print("xlwings error in export_google_sheet_workbook:")
            print(traceback.format_exc())
            raise
    except Exception as e:
        import traceback
        print("xlwings import/open error in export_google_sheet_workbook:")
        print(traceback.format_exc())
        # fallback to openpyxl writer
        try:
            with pd.ExcelWriter(latest_path, engine="openpyxl") as writer:
                for sheet_name, df in sheet_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            print("Latest Excel updated:", latest_path)

            with pd.ExcelWriter(backup_path, engine="openpyxl") as writer:
                for sheet_name, df in sheet_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            print("Backup Excel saved:", backup_path)
        except PermissionError:
            alt = f"reports/AllTeams_Stats_{timestamp}_alt.xlsx"
            with pd.ExcelWriter(alt, engine="openpyxl") as writer:
                for sheet_name, df in sheet_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"Original latest file locked. Saved alternate workbook to {alt}")


KEEPER_COLUMNS = [
    "Club Name", "Date", "Vs Team", "Name of Keeper",
    "Score", "Balls Faced", "Out/Not out", "Catches",
    "Stumps", "Captain Yes\\No", "Match Summary"
]

def rows_to_daily_dataframe(rows):
    import pandas as pd
    df = pd.DataFrame(rows, columns=[
        "club", "date", "vs_team", "keeper", "score", "balls",
        "out_not_out", "catches", "stumps", "captain", "summary", "_match_id"
    ])
    df.columns = KEEPER_COLUMNS + ["_match_id"]
    for col in ["Score", "Balls Faced", "Catches", "Stumps"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    return df


def _write_summary_sheet(wb, df):
    import pandas as pd
    from openpyxl.styles import Font
    num_cols = ["Score", "Balls Faced", "Catches", "Stumps"]
    for c in num_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    bold = Font(bold=True)

    if "Summary" in [s.title for s in wb.sheetnames]:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")

    headers = ["Club Name", "Name of Keeper", "Vs Team", "Date", "Score", "Balls Faced", "Catches", "Stumps", "Out/Not out"]
    for col_idx, col_name in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
        ws.cell(row=1, column=col_idx).font = bold

    df_sorted = df.sort_values("Score", ascending=False).reset_index(drop=True)

    for row_idx, (_, row) in enumerate(df_sorted.iterrows(), 2):
        vals = [
            row.get("Club Name", ""),
            row.get("Name of Keeper", ""),
            row.get("Vs Team", ""),
            row.get("Date", ""),
            int(row["Score"]) if pd.notna(row["Score"]) else 0,
            int(row["Balls Faced"]) if pd.notna(row["Balls Faced"]) else 0,
            int(row["Catches"]) if pd.notna(row["Catches"]) else 0,
            int(row["Stumps"]) if pd.notna(row["Stumps"]) else 0,
            row.get("Out/Not out", ""),
        ]
        for ci, val in enumerate(vals, 1):
            ws.cell(row=row_idx, column=ci, value=val)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 14
    for c in "EFGHI":
        ws.column_dimensions[c].width = 14


def export_append_keeper(rows):
    import pandas as pd
    from openpyxl import load_workbook
    os.makedirs("reports", exist_ok=True)
    master_path = "reports/WicketKeepers_latest.xlsx"

    new_df = rows_to_daily_dataframe(rows)
    if "Date" in new_df.columns:
        new_df["Date"] = new_df["Date"].apply(_normalize_date)

    if os.path.exists(master_path):
        try:
            existing = pd.read_excel(master_path, sheet_name="WicketKeepers", dtype=str)
            if "_match_id" not in existing.columns:
                existing["_match_id"] = ""
            if "Date" in existing.columns:
                existing["Date"] = existing["Date"].apply(_normalize_date)
            combined = pd.concat([existing, new_df], ignore_index=True)
            combined = combined.drop_duplicates(subset=["Name of Keeper", "Date", "Vs Team"], keep="last")
            df_out = combined.copy()
        except Exception:
            df_out = new_df.copy()
    else:
        df_out = new_df.copy()

    if "Date" in df_out.columns:
        df_out["_sort_date"] = pd.to_datetime(df_out["Date"], format="%d %b %Y", errors="coerce")
        df_out = df_out.sort_values("_sort_date", na_position="last").drop(columns=["_sort_date"]).reset_index(drop=True)

    save_cols = [c for c in df_out.columns if c != "_match_id"] + ["_match_id"]
    df_out = df_out[[c for c in save_cols if c in df_out.columns]]

    for attempt in range(2):
        try:
            with pd.ExcelWriter(master_path, engine="openpyxl") as writer:
                df_out.to_excel(writer, sheet_name="WicketKeepers", index=False)
            break
        except PermissionError:
            if attempt == 0:
                print("File locked by Excel. Closing Excel and retrying...")
                close_excel_if_open(master_path)
            else:
                raise

    wb = load_workbook(master_path)
    for s in ["Summary", "Club Summary"]:
        if s in wb.sheetnames:
            del wb[s]
    new_display = new_df.drop(columns=["_match_id"], errors="ignore")
    _write_summary_sheet(wb, new_display.copy())
    wb.save(master_path)
    wb.close()

    print(f"Keeper master Excel updated: {master_path} ({len(new_df)} new rows, total {len(df_out)} unique)")
    return master_path


def export_keeper_dataframe(df):
    os.makedirs("reports", exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    latest_path = "reports/WicketKeepers_latest.xlsx"
    backup_path = f"reports/WicketKeepers_{timestamp}.xlsx"

    try:
        import xlwings as xw
        app = xw.App(visible=False)
        book = None
        try:
            if os.path.exists(latest_path):
                book = app.books.open(latest_path)
            else:
                book = app.books.add()
                book.save(latest_path)
            if "WicketKeepers" in [s.name for s in book.sheets]:
                sht = book.sheets["WicketKeepers"]
            else:
                sht = book.sheets.add("WicketKeepers")
            sht.clear()
            sht.range("A1").options(index=False).value = df
            book.save(latest_path)
            book.close()
            app.quit()
            # Write Summary sheet via openpyxl
            wb = load_workbook(latest_path)
            _write_summary_sheet(wb, df.copy())
            wb.save(latest_path)
            wb.close()
            print("Keeper Excel updated:", latest_path)
        except Exception:
            try:
                if book:
                    book.close()
            except Exception:
                pass
            try:
                app.quit()
            except Exception:
                pass
            raise
    except Exception:
        with pd.ExcelWriter(latest_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="WicketKeepers", index=False)
        wb = load_workbook(latest_path)
        _write_summary_sheet(wb, df.copy())
        wb.save(latest_path)
        wb.close()
        print("Keeper Excel updated:", latest_path)

    shutil.copyfile(latest_path, backup_path)
    print("Keeper backup saved:", backup_path)


def download_google_sheet(sheet_id):
    from pandas import read_excel
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    print("Downloading Google Sheet:", url)
    response = requests.get(url)
    response.raise_for_status()
    temp_path = os.path.join("reports", "google_sheet_download.xlsx")
    with open(temp_path, "wb") as f:
        f.write(response.content)
    workbook = read_excel(temp_path, sheet_name=None)
    return workbook
